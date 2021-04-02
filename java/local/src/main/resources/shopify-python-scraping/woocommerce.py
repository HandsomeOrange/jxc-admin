#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    :   woocommerce.py    
@Contact :   douniwan@douniwan.com
@License :   (C)Copyright 2021-2031, Dou-Ni-Wan

@Modify Time      @Author    @Version    @Description
------------      -------    --------    -----------
2021/3/18 11:58   yqz        1.0         None
"""

# import lib
import json
import re
import threading
import time
import traceback
import openpyxl
import requests
import export_constants as ec

from bs4 import BeautifulSoup


class WooCommerce:

    def __init__(self, site_url):
        self.ses = requests.Session()
        self.ses.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'})
        site_url = site_url.rstrip('/')
        self.shop_url = f'{site_url}/shop'
        self.max_page = self.__get_max_page()
        self.export_headers = ec.shopify_headers

    def __get_max_page(self):
        res = self.ses.get(self.shop_url)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, features='lxml')
        max_page = 0
        for s in soup.find_all('a', class_='page-number'):
            if s.string:
                number_str = re.sub('[^0-9]', '', s.string)
                max_page = max_page if max_page > int(number_str) else int(number_str)
        return max_page

    def __get_page_data(self, page_list, data_list):
        for page_index in page_list:
            page_response = self.ses.get(f'{self.shop_url}/page/{page_index}/')
            page_response.raise_for_status()
            page_soup = BeautifulSoup(page_response.text, features='lxml')
            # for div in page_soup.find_all('p', class_='product-title'):
            #     print(div)

            for p in page_soup.select('p.product-title a'):
                data_list.append({'title': p.string, 'href': p['href']})

    def __get_product_detail(self, data_list, item_list):
        for d in data_list:
            product_detail = self.ses.get(d['href'])
            product_detail.raise_for_status()
            detail_soup = BeautifulSoup(product_detail.text, features='lxml')
            tag_list = []
            for tag_a in detail_soup.select('nav.woocommerce-breadcrumb.breadcrumbs.uppercase a'):
                tag_list.append(tag_a.string)
            option_list = []
            for tag_label in detail_soup.select('table.variations label'):
                option_list.append((tag_label['for'], tag_label.string))
            option_list.sort(key=lambda k: k[1])
            option_key_0, option_name_0 = (f'attribute_{option_list[0][0]}', option_list[0][1]) if len(
                option_list) > 0 else ('x', '')
            option_key_1, option_name_1 = (f'attribute_{option_list[1][0]}', option_list[1][1]) if len(
                option_list) > 1 else ('x', '')
            option_key_2, option_name_2 = (f'attribute_{option_list[2][0]}', option_list[2][1]) if len(
                option_list) > 2 else ('x', '')
            form = detail_soup.find(class_='variations_form cart')
            product_tittle = d['title']
            handle = re.sub('[^0-9a-zA-Z]', '-', product_tittle).lower()

            m_product = {'0': handle, '1': product_tittle, '2': '', '5': ','.join(tag_list),
                         '6': 'TRUE',
                         '7': option_name_0, '8': option_name_0, '9': option_name_1,
                         '10': option_name_1, '11': option_name_2, '12': option_name_2}

            # {'5': product_tittle, '6': product_tittle, '8': 'N', '9': 'Y', '10': 'N', '11': '0', '12': 'N',
            #  '13': '2', '15': ','.join(tag_list), '18': 'M',
            #  '19': option_name_0, '20': option_name_1, '21': option_name_2,
            #  '47': form['data-product_id'] if form.has_attr('data-product_id') else ''}
            desc = ''
            description_tag = detail_soup.select_one('div#tab-description')
            if description_tag:
                desc = description_tag.text
            p_products = []
            if form and form.has_attr('data-product_variations'):
                detail_json = json.loads(form['data-product_variations'])
                for ds in detail_json:
                    # product_item = {'0': ds['variation_id'], '1': d['title'], '15': ','.join(tag_list), '18': 'P'}
                    product_item = {'0': handle, '1': product_tittle, '2': desc, '5': ','.join(tag_list),
                                    '6': 'TRUE',
                                    '7': option_name_0, '8': option_name_0, '9': option_name_1,
                                    '10': option_name_1, '11': option_name_2, '12': option_name_2}
                    attr = ds['attributes']
                    if attr:
                        product_item['8'] = attr[option_key_0] if option_key_0 in attr else ''
                        product_item['10'] = attr[option_key_1] if option_key_1 in attr else ''
                        product_item['12'] = attr[option_key_2] if option_key_2 in attr else ''
                    product_item['19'] = ds['display_price']
                    product_item['20'] = ds['display_regular_price']
                    product_item['13'] = ds['sku']
                    product_item['14'] = ds['weight']
                    product_item['16'] = 0
                    product_item['17'] = 'continue'
                    product_item['18'] = 'manual'
                    product_item['21'] = 'TRUE'
                    product_item['22'] = 'manual'
                    product_item['24'] = ds['image']['src'] if ds['image'] and 'src' in ds['image'] else ''
                    product_item['26'] = ds['image']['alt'] if ds['image'] and 'alt' in ds['image'] else ''
                    product_item['27'] = 'FALSE'
                    product_item['28'] = product_tittle
                    product_item['29'] = product_tittle
                    product_item['43'] = product_item['24']
                    product_item['47'] = ds['variation_id']
                    product_item['48'] = 'P'
                    p_products.append(product_item)
            # if not len(p_products):
            #     m_product['18'] = 'S'
            # item_list.append(m_product)
            item_list += p_products

    def __export_all_data(self):
        result = []
        try:
            ds = []
            thread_cnt = 60
            # //self.max_page
            total_page_index = [i for i in range(1, 10 + 1)]
            page_list = [total_page_index[i::thread_cnt] for i in range(thread_cnt)]
            page_threads = []
            for thread_id in range(len(page_list)):
                page_threads.append(
                    threading.Thread(target=self.__get_page_data, args=(page_list[thread_id], ds,)))
            for t in page_threads:
                t.setDaemon(True)
                t.start()
            for t in page_threads:
                t.join()
            print(f'get data list end {len(ds)}')

            little_list = [ds[i::thread_cnt] for i in range(thread_cnt)]
            threads = []
            for thread_id in range(len(little_list)):
                threads.append(
                    threading.Thread(target=self.__get_product_detail, args=(little_list[thread_id], result,)))
            for t in threads:
                t.setDaemon(True)
                t.start()
            for t in threads:
                t.join()

            # for i in result:
            #     print(i)
        except Exception as e:
            print("save to db error ", e)
            traceback.print_exc()
        return result

    def export_to_excel(self, file_name):
        all_data = self.__export_all_data()
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet('export', 0)
        # item_dict.has_pic,item_dict.has_video, item_dict.save_path,
        # template_wb = openpyxl.load_workbook('product_import_template_shoplazza.xlsx', read_only=True)
        # template_sheet = template_wb.worksheets[0]
        # sheet.append()
        for i in range(len(self.export_headers)):
            sheet.cell(1, i + 1, self.export_headers[i])
        row = 1
        for d_item in all_data:
            row += 1
            for i in range(len(self.export_headers)):
                sheet.cell(row, i + 1, d_item[str(i)] if str(i) in d_item else '')
        # output_file_name = f'{save_xlsx_path}\\result.xlsx'
        wb.save(rf'export\{file_name}')
        wb.close()
        print(f'output {file_name} success >>>>>> ')

    def test(self):
        self.__get_product_detail([{'href': 'https://gifnest.com/product/kb-jd13-sneaker/', 'title': 'xxx'}], [])


if __name__ == '__main__':
    wo = WooCommerce('https://gifnest.com/')
    # wo.test()
    now = time.strftime('%Y%m%d%H%M%S')
    wo.export_to_excel(f'export_{now}.xlsx')
